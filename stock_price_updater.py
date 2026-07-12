'''
stock_price_updater.py

'''
from datetime import datetime
import pandas as pd
import twstock
import twstock.stock
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

XLSX_PATH = 'stock_history.xlsx'
SHEET_NAME = 'Stock'
GREEN_COLOR = "7AD694"
RED_COLOR = "F28E86"

def _patch_twstock_extra_columns():
    """
    twstock 1.4.0 expects 9 TWSE/TPEX data columns. TWSE currently returns an
    extra trailing note column, so trim any new trailing columns before twstock
    builds its fixed Data tuple.
    """
    for fetcher_name in ("TWSEFetcher", "TPEXFetcher"):
        fetcher = getattr(twstock.stock, fetcher_name, None)
        if fetcher is None or getattr(fetcher, "_stock_analyse_patched", False):
            continue

        original_make_datatuple = fetcher._make_datatuple

        def make_datatuple(self, data, original_make_datatuple=original_make_datatuple):
            return original_make_datatuple(self, list(data[:9]))

        fetcher._make_datatuple = make_datatuple
        fetcher._stock_analyse_patched = True


def fetch_latest_stock_price(stock_tar):
    """
    Fetch the latest trading data for a given stock,
    starting from the beginning of the current year to the last full month.

    Args:
        stock_tar (str): The stock code to query.

    Returns:
        dict: A dictionary containing the latest stock's trading data.
    """

    # Get the current date
    now = datetime.now()

    # Set start year and the last full month
    start_year = now.year
    start_month = now.month - 1 if now.month > 1 else 12

    # Patch only when the twstock fetch path is used.
    _patch_twstock_extra_columns()

    # Create a Stock object
    stock = twstock.Stock(stock_tar)

    # Fetch daily trading data from the start of the current year
    # to the last full month
    target_price = stock.fetch_from(start_year, start_month)

    # Set headers for the collected data
    name_attribute = [
        'Date',  # 日期
        'Capacity',  # 總成交股數
        'Turnover',  # 總成交金額(Volume)
        'Open',
        'High',
        'Low',
        'Close',
        'Change',  # 漲跌幅
        'Transaction'  # 成交量
    ]

    df = pd.DataFrame(columns=name_attribute, data=target_price)

    # Access the latest record (last row in DataFrame)
    return df.iloc[-1].to_dict()

def get_col(header, col_name):
    if not header.get(col_name):
        print(f"{col_name} not found in the header.")
        return None
    else:
        return header.get(col_name)

if __name__ == "__main__":
    # Load the workbook
    wb = load_workbook(filename=XLSX_PATH)
    sheet = wb[SHEET_NAME]

    # Find the headers
    header = {cell.value: cell.column_letter for cell in sheet[1]}

    # Input column
    col_stock_index    = get_col(header, 'Stock Index')
    col_amount         = get_col(header, 'Amount')
    col_average_cost   = get_col(header, 'Average Cost')
    # Output column
    col_current_price  = get_col(header, 'Current Price')
    col_current_profit = get_col(header, 'Current Profit')
    col_profit_percent = get_col(header, 'Profit Percentage')
    col_current_value  = get_col(header, 'Current Value')
    col_sum_value      = get_col(header, 'Sum Value')

    sum_value = 0
    # Iterate over the rows and process the stock indices
    for row in range(2, sheet.max_row + 1):
        # Get stock index
        stock_tar = str(sheet[f'{col_stock_index}{row}'].value)
        if stock_tar == 'None':
            continue

        # Fetch stock price from the Internet
        try:
            latest_data = fetch_latest_stock_price(stock_tar)
            close_price = latest_data['Close']  # Get the 'Close' value
        except Exception as e:
            close_price = sheet[f'{col_current_price}{row}'].value
            print(f"Error fetching data for stock index {stock_tar}: {e}")
        else:
            # Update 'Current Price'
            sheet[f'{col_current_price}{row}'] = close_price
            print(f"Updated stock index {stock_tar} with " +
                f"latest close price {close_price}")

        # Get cost and amount
        cost   = sheet[f'{col_average_cost}{row}'].value
        amount = sheet[f'{col_amount}{row}'].value

        # Update 'Current Profit'
        sheet[f'{col_current_profit}{row}'] = (close_price - cost) * amount

        # Update 'Profit Percentage'
        sheet[f'{col_profit_percent}{row}'] = (close_price - cost) / cost

        # Update 'Current Value'
        sheet[f'{col_current_value}{row}'] = close_price * amount

        sum_value += sheet[f'{col_current_value}{row}'].value

        # Apply background color based on the profit values
        if sheet[f'{col_current_profit}{row}'].value < 0:
            # Negative profit
            sheet[f'{col_current_profit}{row}'].fill = PatternFill(start_color=GREEN_COLOR,
                                                                   end_color=GREEN_COLOR,
                                                                   fill_type="solid")
            sheet[f'{col_profit_percent}{row}'].fill = PatternFill(start_color=GREEN_COLOR,
                                                                   end_color=GREEN_COLOR,
                                                                   fill_type="solid")
        else:
            # Positive profit
            sheet[f'{col_current_profit}{row}'].fill = PatternFill(start_color=RED_COLOR,
                                                                   end_color=RED_COLOR,
                                                                   fill_type="solid")
            sheet[f'{col_profit_percent}{row}'].fill = PatternFill(start_color=RED_COLOR,
                                                                   end_color=RED_COLOR,
                                                                   fill_type="solid")


    # Update sum value
    sheet[f'{col_sum_value}{2}'] = sum_value

    # Save the workbook
    try:
        wb.save(filename=XLSX_PATH)
        print("Workbook saved successfully.")
    except PermissionError:
        print("Permission denied: The file might be open in " +
                "another program or you might not have the " +
                "necessary permissions to write to the file.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
